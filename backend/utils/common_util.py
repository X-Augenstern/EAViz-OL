from pandas import DataFrame
from io import BytesIO
from os import path
from re import sub
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.engine.row import Row
from typing import List
from config.env import CachePathConfig


def worship():
    print("""
////////////////////////////////////////////////////////////////////
//                          _ooOoo_                               //
//                         o8888888o                              //
//                         88" . "88                              //
//                         (| ^_^ |)                              //
//                         O\  =  /O                              //
//                      ____/`---'\____                           //
//                    .'  \\|     |//  `.                         //
//                   /  \\|||  :  |||//  \                        //
//                  /  _||||| -:- |||||-  \                       //
//                  |   | \\\  -  /// |   |                       //
//                  | \_|  ''\---/''  |   |                       //
//                  \  .-\__  `-`  ___/-. /                       //
//                ___`. .'  /--.--\  `. . ___                     //
//              ."" '<  `.___\_<|>_/___.'  >'"".                  //
//            | | :  `- \`.;`\ _ /`;.`/ - ` : | |                 //
//            \  \ `-.   \_ __\ /__ _/   .-` /  /                 //
//      ========`-.____`-.___\_____/___.-`____.-'========         //
//                           `=---='                              //
//      ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^        //
//              佛祖保佑       永不宕机       永无BUG                 //
////////////////////////////////////////////////////////////////////
    """)


class CamelCaseUtil:
    """
    小驼峰形式(camelCase)与下划线形式(snake_case)互相转换工具方法
    """
    @classmethod
    def camel_to_snake(cls, camel_str):
        """
        小驼峰形式字符串(camelCase)转换为下划线形式字符串(snake_case)
        :param camel_str: 小驼峰形式字符串
        :return: 下划线形式字符串
        """
        # sub(count = 0)：替换字符串中所有符合正则表达式的部分，而不仅仅是第一次匹配的部分。它会遍历整个字符串，对每个匹配的部分进行替换操作
        # 在整个匹配过程中，正则表达式会尝试在字符串中找到一个位置，使得整个表达式都能成功匹配
        # 如字符串 "123abc"，整个表达式 (\d+)([a-z]+) 在字符串 "123abc" 中从头到尾都能找到匹配，其中第一组匹配 "123"，紧接着第二组匹配 "abc"
        # 所以，第一组和第二组的匹配是基于整个表达式对字符串进行扫描和匹配的结果，并非是“去除”前一个组的匹配结果后进行的。整个匹配过程是连续和整体的
        # 当找到一个匹配项，该函数会将匹配的部分替换掉，然后从这个匹配项后的下一个字符继续寻找下一个匹配

        # (.) 是第一个捕获组，匹配任意单个字符                                      l
        # ([A-Z][a-z]+) 是第二个捕获组，匹配一个大写字母后跟至少一个小写字母            Case
        # r'\1_\2' 是替换模式
        # \1 指代第一个捕获组的内容，即字符
        # _ 是字面意义上的下划线字符，用于替换驼峰命名中的大写字母前的位置                l_Case
        # \2 指代第二个捕获组的内容，即一个完整的词的起始部分（如“Camel”中的“Camel”）

        # 处理剩余的驼峰命名情况，即小写字母或数字后紧跟一个大写字母的情况
        # ([a-z0-9]) 是第一个捕获组，匹配一个小写字母或数字                           e
        # ([A-Z]) 是第二个捕获组，匹配一个大写字母                                   E
        #                                                                      e_E

        # 在大写字母前添加一个下划线，然后将整个字符串转为小写
        words = sub('(.)([A-Z][a-z]+)', r'\1_\2', camel_str)  # "CamelCaseExample" ——> "Camel_CaseExample"
        return sub('([a-z0-9])([A-Z])', r'\1_\2', words).lower()  # "Camel_CaseExample" ——> "camel_case_example"

    @classmethod
    def snake_to_camel(cls, snake_str):
        """
        下划线形式字符串(snake_case)转换为小驼峰形式字符串(camelCase)
        :param snake_str: 下划线形式字符串
        :return: 小驼峰形式字符串
        """
        # 分割字符串
        words = snake_str.split('_')
        # capitalize()：将字符串的首字母转换为大写，其余字母转换为小写
        # 小驼峰命名，第一个词首字母小写，其余词首字母大写
        return words[0] + ''.join(word.capitalize() for word in words[1:])

    @classmethod
    def transform_result_rewrite(cls, result):
        """
        针对不同类型将下划线形式(snake_case)批量转换为小驼峰形式(camelCase)方法
        :param result: 输入数据
        :return: 小驼峰形式结果
        """
        if result is None:
            return result
        elif isinstance(result, dict):
            return {cls.snake_to_camel(k): cls.snake_to_camel(v) for k, v in result.items()}
        elif isinstance(result, list):
            return [cls.transform_result(item) for item in result]
        elif isinstance(result, Row):
            return {cls.snake_to_camel(c.name): getattr(result, c.name) for c in result.__table__.columns}
        else:
            return result

    @classmethod
    def transform_result(cls, result):
        """
        针对不同类型将下划线形式(snake_case)批量转换为小驼峰形式(camelCase)方法
        :param result: 输入数据
        :return: 小驼峰形式结果
        """
        if result is None:
            return result
        # 如果是字典，直接转换键
        elif isinstance(result, dict):
            return {cls.snake_to_camel(k): v for k, v in result.items()}
        # 如果是一组字典或其他类型的列表，遍历列表进行转换
        # 如果列表内的元素是 dict/Row ，再递归调用此函数，否则先创建字典，其中包含模型的列名作为键，相应的属性值作为值，再递归调用此函数
        elif isinstance(result, list):
            return [cls.transform_result(row) if isinstance(row, (dict, Row)) else (
                cls.transform_result({c.name: getattr(row, c.name) for c in row.__table__.columns}) if row else row) for
                    row in result]
        # 如果是sqlalchemy的Row实例，遍历Row进行转换
        elif isinstance(result, Row):
            return [cls.transform_result(row) if isinstance(row, dict) else (
                cls.transform_result({c.name: getattr(row, c.name) for c in row.__table__.columns}) if row else row) for
                    row in result]
        # 如果是其他类型，如模型实例，先转换为字典
        else:
            # result.__table__.columns 访问该模型的所有列的元数据。这是 SQLAlchemy 中模型类的一个常见用法，用于获取模型的列信息
            # c.name：对于模型中的每一列 c，c.name 表示列的名称

            # getattr(result, c.name)：这个函数调用获取 result 对象的属性，属性名为 c.name。这里基本上是在获取每一列的值
            # Get a named attribute from an object; getattr(x, 'y') is equivalent to x.y
            # getattr 是一个内置函数，用于动态地从对象中获取属性。第一个参数是对象，第二个参数是字符串形式的属性名
            # 如果属性名是动态确定的，或者在编写代码时不知道具体的属性名，getattr 特别有用
            # 例如，在遍历数据库模型的列时，列名是动态的，因此使用 getattr 可以根据列名字符串获取相应的属性值
            # 使用 getattr 还可以方便地设置默认值。如果属性不存在，可以指定 getattr 的第三个参数作为默认返回值，以避免抛出 AttributeError
            # result.c.name 这种方式是直接属性访问，其中 c 和 name 都是硬编码的属性名，适用于属性名已知且不变的情况

            # 从一个 ORM 模型实例中提取数据，将其转换为一个键为列名、值为相应列值的字典，然后通过一个类方法对这个字典进行进一步的处理或格式化
            # {'id': 1, 'name': 'Alice', 'email': 'alice@example.com'}
            return cls.transform_result({c.name: getattr(result, c.name) for c in result.__table__.columns})


def bytes2human(n, format_str="%(value).1f%(symbol)s"):  # %(value).1f 表示小数点后保留一位的浮点数 %(symbol)s 表示字符串格式的单位符号
    """Used by various scripts. See:
    http://goo.gl/zeJZl

    '>>> bytes2human(10000)'
    '9.8K'
    '>>> bytes2human(100001221)'
    '95.4M'
    """
    symbols = ('B', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB')
    prefix = {}
    for i, s in enumerate(symbols[1:]):
        prefix[s] = 1 << (i + 1) * 10
    for symbol in reversed(symbols[1:]):
        if n >= prefix[symbol]:
            value = float(n) / prefix[symbol]
            # locals() 将当前作用域中定义的所有局部变量（例如 symbol 和 value）作为一个字典传递给字符串格式化操作
            # 当在字符串中使用 %(variable_name)s 或 %(variable_name)f 这样的格式化占位符时，
            # Python 会查找一个字典中与 variable_name 对应的键，并替换为其值
            # 这样做允许 format_str 动态地访问这些变量并进行格式化，而无需显式指定每个变量
            # 假设某个时刻 value 变量的值是 1024.0，symbol 的值是 'KB'，并且 format_str 被定义为 "%(value).1f%(symbol)s"
            # 通过使用 locals()，format_str 中的 %(value).1f 和 %(symbol)s 将被替换为 1024.0 和 KB，结果字符串就会是 "1024.0KB"
            # 简化了字符串格式化过程，尤其是在需要格式化多个变量时，使用 locals() 可以避免单独传递每一个变量
            return format_str % locals()
    return format_str % dict(symbol=symbols[0], value=n)  # 'B'


def bytes2file_response(bytes_info):
    """
    包含 yield 关键字时，这个函数被称为生成器（generator）。生成器是一种特殊类型的迭代器，它允许函数生成一个数据序列，而不是一次性返回所有数据。

    每次产生一个值后会暂停其状态（包括局部变量和执行位置），直到下一次从该生成器请求数据时再继续执行。这种机制使得生成器非常适合处理大数据流或需要懒加载（按需加载）的场景。

    假设 bytes_info 是一些二进制数据，这个生成器可以被用于按部分传输数据，例如在 Web 应用中流式传输文件内容，允许用户下载大文件而不必将整个文件一次性加载到内存中。

    Notes：
    目前并没有实际地将数据分块发送，而是将整个字节数据作为一个整体发送。这个函数只是简单地将整个 bytes_info 对象作为一个块进行返回，并没有进行分块处理。

    这样的话，虽然它是一个生成器，但实际上仍然是一次性加载和发送所有数据，没有实现逐步发送数据的效果。

    实现真正的分块发送：分割数据，并逐块返回
    chunk_size = 8192 # 默认块大小为 8192 字节（8KB）
    for i in range(0, len(bytes_info), chunk_size):
        yield bytes_info[i:i + chunk_size]
    """
    yield bytes_info


def export_list2excel(list_data: List):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据
    :param list_data: 数据列表
    :return: 字典信息对应excel的二进制数据
    """
    df = DataFrame(list_data)
    binary_data = BytesIO()
    # index=False 表示在生成的 Excel 文件中不包括行索引（即不添加额外的索引列）
    df.to_excel(binary_data, index=False, engine='openpyxl')  # 将df写入内存中的二进制流中
    binary_data = binary_data.getvalue()  # 从 BytesIO 对象中获取完整的二进制内容

    return binary_data


def get_excel_template(header_list: List, selector_header_list: List, option_list: List[dict]):
    """
    工具方法：将需要导出的list数据转化为对应excel的二进制数据
    :param header_list: 表头数据列表
    :param selector_header_list: 需要设置为选择器格式的表头数据列表
    :param option_list: 选择器格式的表头预设的选项列表
    :return: 模板excel的二进制数据
    """
    # 创建Excel工作簿
    wb = Workbook()
    # 选择默认的活动工作表（属性：返回的是当前活动的工作表对象，可以直接对其进行操作，比如添加数据、设置格式等）
    ws = wb.active

    # 设置表头文字
    headers = header_list

    # 设置表头背景样式为灰色，前景色为白色
    # 颜色代码通常为六位十六进制数，每两位代表红（R）、绿（G）、蓝（B）三原色的一个分量。每个分量的值范围是从00（最暗）到FF（最亮）
    # 转换为十进制 171 表示红、绿、蓝三种颜色以相等的中等亮度混合，生成了一种中等深浅的灰色
    header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")

    # 将表头写入第一行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        # 设置列宽度为16 / ws.column_dimensions[get_column_letter(col_num)].width = 12
        ws.column_dimensions[chr(64 + col_num)].width = 12  # chr(64 + col_num) 将列索引转化为excel中的ABCD...
        # 设置水平居中对齐
        cell.alignment = Alignment(horizontal='center')

    # 设置选择器的预设选项
    options = option_list

    # 获取selector_header的字母索引 = 在headers内的idx+1
    for selector_header in selector_header_list:
        column_selector_header_index = headers.index(selector_header) + 1
        # 对selector_header_list中指定的每列，根据option_list中提供的选项设置添加下拉选择功能，让用户只能从预设的选项中选择数据
        # 创建数据有效性规则
        header_option = []
        for option in options:
            if option.get(selector_header):  # 如果当前列标题在字典中存在，就将这些选项赋值给header_option
                header_option = option.get(selector_header)
        # 验证规则允许的值必须来自一个预定义的列表 eg：header_option = ["销售", "市场", "研发", "人力资源"]
        # formula1属性设置将所有选项拼接成一个逗号分隔的字符串，并包裹在双引号内 eg：'"销售,市场,研发,人力资源"'
        dv = DataValidation(type="list", formula1=f'"{",".join(header_option)}"')
        # get_column_letter 将列索引转换为对应的列字母
        # 设置数据有效性规则的起始单元格和结束单元格
        # dv.add(...)设置这个数据验证规则应用于从第二行开始到最大行（1048576行）的当前列。这样做是为了不覆盖第一行的表头
        # eg：dv.add('B2:B1048576') 假设部门列是B列，从第二行开始应用（第一行通常是表头）
        dv.add(
            f'{get_column_letter(column_selector_header_index)}2:{get_column_letter(column_selector_header_index)}1048576')
        # 添加数据有效性规则到工作表，这样被指定的列就会具备下拉选择功能，用户编辑单元格时会显示一个下拉列表，从中选择数据
        ws.add_data_validation(dv)

    # 保存Excel文件为字节类型的数据
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    # 读取字节数据
    excel_data = file.getvalue()

    return excel_data


def get_filepath_from_url(url: str):
    """
    工具方法：根据请求参数获取文件路径
    :param url: 请求参数中的url参数
    :return: 文件路径

    eg: http://example.com/getfile?task_id=123&file_name=myfile.txt&task_path=/data
    """
    file_info = url.split("?")[1].split("&")  # 解析查询参数：["task_id=123", "file_name=myfile.txt", "task_path=/data"]
    task_id = file_info[0].split("=")[1]
    file_name = file_info[1].split("=")[1]
    task_path = file_info[2].split("=")[1]
    filepath = path.join(CachePathConfig.PATH, task_path, task_id, file_name)

    return filepath
